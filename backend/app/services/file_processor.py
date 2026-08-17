"""
Parse uploaded spreadsheet files (xlsx / xlsm / csv) and persist Sheet +
SheetRow records.  Called synchronously from the upload endpoint; swap for
an arq task in a later phase when async processing is required.

Row cap: MAX_ROWS_PER_SHEET — files exceeding this are truncated (the cap
protects DB insert time and avoids memory spikes on large uploads).
"""

from __future__ import annotations

import io
from pathlib import Path
from typing import Any

import openpyxl
import pandas as pd
from sqlalchemy import insert, update
from sqlalchemy.ext.asyncio import AsyncSession
from sqlalchemy.orm import selectinload

from app.models.file import File, FileStatus
from app.models.sheet import Sheet, SheetRow

MAX_ROWS_PER_SHEET = 100_000
SUPPORTED_EXTENSIONS = {".xlsx", ".xlsm", ".csv"}


def _dtype_label(value: Any) -> str:
    if isinstance(value, bool):
        return "boolean"
    if isinstance(value, (int, float)):
        return "number"
    if hasattr(value, "isoformat"):
        return "date"
    return "text"


def _to_jsonable(value: Any) -> Any:
    if value is None:
        return None
    if isinstance(value, bool):
        return value
    if isinstance(value, (int, float)):
        return value
    if hasattr(value, "isoformat"):
        return value.isoformat()
    return str(value)


def _parse_excel(file_bytes: bytes) -> list[tuple[str, list[dict], list[dict]]]:
    wb = openpyxl.load_workbook(io.BytesIO(file_bytes), read_only=True, data_only=True)
    results: list[tuple[str, list[dict], list[dict]]] = []

    for sheet_name in wb.sheetnames:
        ws = wb[sheet_name]
        rows_iter = ws.iter_rows(values_only=True)

        try:
            header_row = next(rows_iter)
        except StopIteration:
            results.append((sheet_name, [], []))
            continue

        headers = [str(h) if h is not None else f"col_{i}" for i, h in enumerate(header_row)]

        # Deduplicate headers
        seen: dict[str, int] = {}
        unique_headers = []
        for h in headers:
            if h in seen:
                seen[h] += 1
                unique_headers.append(f"{h}_{seen[h]}")
            else:
                seen[h] = 0
                unique_headers.append(h)
        headers = unique_headers

        dtype_map: dict[str, str] = {}
        rows_data: list[dict] = []

        for raw in rows_iter:
            if len(rows_data) >= MAX_ROWS_PER_SHEET:
                break
            row_dict = {
                headers[i]: _to_jsonable(v)
                for i, v in enumerate(raw)
                if i < len(headers)
            }
            if not dtype_map:
                dtype_map = {headers[i]: _dtype_label(v) for i, v in enumerate(raw) if i < len(headers)}
            rows_data.append(row_dict)

        columns_meta = [
            {"name": h, "dtype": dtype_map.get(h, "text"), "index": i, "width": 120}
            for i, h in enumerate(headers)
        ]
        results.append((sheet_name, columns_meta, rows_data))

    wb.close()
    return results


def _parse_csv(file_bytes: bytes) -> list[tuple[str, list[dict], list[dict]]]:
    try:
        text = file_bytes.decode("utf-8-sig")
    except UnicodeDecodeError:
        text = file_bytes.decode("latin-1")

    df = pd.read_csv(io.StringIO(text), nrows=MAX_ROWS_PER_SHEET, dtype=str)
    df = df.where(pd.notna(df), None)

    # Deduplicate column names
    cols: list[str] = []
    seen: dict[str, int] = {}
    for c in df.columns.tolist():
        c = str(c)
        if c in seen:
            seen[c] += 1
            cols.append(f"{c}_{seen[c]}")
        else:
            seen[c] = 0
            cols.append(c)
    df.columns = cols  # type: ignore[assignment]

    columns_meta = [
        {"name": c, "dtype": "text", "index": i, "width": 120}
        for i, c in enumerate(cols)
    ]
    rows_data = df.to_dict(orient="records")
    return [("Sheet1", columns_meta, rows_data)]


async def process_file(
    file_bytes: bytes,
    filename: str,
    file_record: File,
    db: AsyncSession,
) -> File:
    """Parse *file_bytes* and persist Sheet + SheetRow records.

    Updates *file_record* status to ready/failed and commits.  Returns the
    refreshed File ORM object (with sheets eagerly loaded).
    """
    from sqlalchemy import select

    ext = Path(filename).suffix.lower()

    # ── Parse (CPU-bound, no I/O) ──────────────────────────────────────────────
    try:
        if ext in (".xlsx", ".xlsm"):
            sheets_data = _parse_excel(file_bytes)
        elif ext == ".csv":
            sheets_data = _parse_csv(file_bytes)
        else:
            raise ValueError(f"Unsupported file type: {ext!r}")
    except Exception as exc:
        file_record.status = FileStatus.failed
        file_record.error_message = str(exc)[:500]
        await db.commit()
        return file_record

    # ── Mark as processing ────────────────────────────────────────────────────
    file_record.status = FileStatus.processing
    await db.commit()

    # ── Insert sheets + rows ──────────────────────────────────────────────────
    try:
        total_rows = 0
        for idx, (sheet_name, columns_meta, rows_data) in enumerate(sheets_data):
            sheet = Sheet(
                file_id=file_record.id,
                name=sheet_name,
                sheet_index=idx,
                row_count=len(rows_data),
                col_count=len(columns_meta),
                columns=columns_meta,
            )
            db.add(sheet)
            await db.flush()  # materialise sheet.id

            if rows_data:
                await db.execute(
                    insert(SheetRow),
                    [
                        {"sheet_id": sheet.id, "row_index": i, "data": row}
                        for i, row in enumerate(rows_data)
                    ],
                )
            total_rows += len(rows_data)

        file_record.status = FileStatus.ready
        file_record.sheet_count = len(sheets_data)
        file_record.total_rows = total_rows
        await db.commit()

    except Exception as exc:
        await db.rollback()
        await db.execute(
            update(File)
            .where(File.id == file_record.id)
            .values(status=FileStatus.failed, error_message=str(exc)[:500])
        )
        await db.commit()
        file_record.status = FileStatus.failed
        return file_record

    # ── Reload with sheets ────────────────────────────────────────────────────
    result = await db.execute(
        select(File)
        .where(File.id == file_record.id)
        .options(selectinload(File.sheets))
    )
    return result.scalar_one()
